from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import pandas as pd
import csv

SETTLEMENT_DIR = Path(r'C:\Users\jnpei\Documents\stock-chart\data\settlement')
OUTPUT_FILE    = SETTLEMENT_DIR / 'earnings.csv'

# JPX xlsx の有効な決算種別
VALID_TYPES = {'本決算', '第１四半期', '第２四半期', '第３四半期', '第４四半期'}

# kessan_pro の決算期間 → 統一種別へのマッピング
PERIOD_MAP = {
    '通期':     '本決算',
    '第1四半期': '第１四半期',
    '第2四半期': '第２四半期',
    '第3四半期': '第３四半期',
}

# ---- ユーティリティ ----------------------------------------------------------

def normalize_sc(sc):
    """SC を 4 桁ゼロ埋め文字列に正規化する。"""
    if sc is None:
        return None
    if isinstance(sc, (int, float)):
        return str(int(sc)).zfill(4)
    return str(sc).strip().zfill(4)


def normalize_date(d):
    """日付を YYYY/MM/DD 形式に正規化する。"""
    if isinstance(d, datetime):
        return d.strftime('%Y/%m/%d')
    if isinstance(d, str):
        d = d.strip()
        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日'):
            try:
                return datetime.strptime(d, fmt).strftime('%Y/%m/%d')
            except ValueError:
                continue
    return None


# ---- JPX xlsx 読み込み -------------------------------------------------------

def load_jpx_xlsx(filepaths: list[Path]) -> pd.DataFrame:
    """JPX 形式の xlsx（kessan[0-9]*.xlsx）を読み込み DataFrame を返す。

    フォーマット:
      Row 0〜3: タイトル等, Row 4: 列名行 → 先頭 5 行スキップ
      row[0]=発表日, row[1]=SC, row[7]=決算種別
    """
    seen  = set()
    rows  = []

    for xlsx_path in sorted(filepaths):
        print(f'[JPX] 読み込み中: {xlsx_path.name}')
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as e:
            print(f'  スキップ（読み込みエラー）: {e}')
            continue

        for sheet_name in wb.sheetnames:
            ws       = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            for _ in range(5):          # 先頭 5 行スキップ
                try:
                    next(rows_iter)
                except StopIteration:
                    break

            for row in rows_iter:
                if row is None or len(row) < 8:
                    continue

                raw_type = row[7]
                if raw_type not in VALID_TYPES:
                    continue

                date_str = normalize_date(row[0])
                if date_str is None:
                    continue

                sc_str = normalize_sc(row[1])
                if not sc_str:
                    continue

                key = (sc_str, date_str)
                if key in seen:
                    continue
                seen.add(key)
                rows.append({'SC': sc_str, 'date': date_str, 'type': raw_type})

        wb.close()

    return pd.DataFrame(rows, columns=['SC', 'date', 'type'])


# ---- kessan_pro xlsx 読み込み ------------------------------------------------

def load_kessan_pro(filepath: Path) -> pd.DataFrame:
    """kessan_pro 形式の xlsx（kessan_pro_*.xlsx）を読み込み DataFrame を返す。

    連結優先・個別補完方式:
      - 連結データが存在する SC → 連結行のみ取り込む
      - 連結データが存在しない SC → 個別行を取り込む（個別企業を補完）

    フォーマット（0-indexed 列番号）:
      row[0]  : SC コード
      row[3]  : 連結個別（'連結' / '個別'）
      row[5]  : 決算期間（PERIOD_MAP でフィルタ）
      row[21] : 情報公開又は更新日（発表日として使用）
    """
    print(f'[Pro]  読み込み中: {filepath.name}')

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f'  スキップ（読み込みエラー）: {e}')
        return pd.DataFrame(columns=['SC', 'date', 'type'])

    ws   = wb.active
    all_rows = list(ws.iter_rows(values_only=True))[1:]  # ヘッダー行スキップ
    wb.close()

    # Step1: 連結データが存在する SC のセットを先に収集
    renketsu_scs = set()
    for row in all_rows:
        if row is None or len(row) < 22:
            continue
        sc = normalize_sc(row[0])
        if sc and row[3] == '連結':
            renketsu_scs.add(sc)

    # Step2: 連結優先・個別補完でデータ取り込み
    seen      = set()
    records   = []
    n_kobetsu = 0

    for row in all_rows:
        if row is None or len(row) < 22:
            continue

        sc       = normalize_sc(row[0])
        renketsu = row[3]   # 連結個別
        period   = row[5]   # 決算期間
        pub_date = row[21]  # 情報公開又は更新日

        if not sc:
            continue
        if period not in PERIOD_MAP:
            continue

        # 連結優先・個別補完の判定
        is_kobetsu_sc = sc not in renketsu_scs
        if sc in renketsu_scs:
            if renketsu != '連結':      # 連結SCは連結行のみ
                continue
        else:
            if renketsu != '個別':      # 連結なしSCは個別行を補完
                continue

        date_str = normalize_date(pub_date)
        if date_str is None:
            continue

        type_str = PERIOD_MAP[period]
        key      = (sc, date_str)
        if key in seen:
            continue
        seen.add(key)
        if is_kobetsu_sc:
            n_kobetsu += 1
        records.append({'SC': sc, 'date': date_str, 'type': type_str})

    print(f'  連結SCセット   : {len(renketsu_scs):,} 社')
    print(f'  個別補完レコード: {n_kobetsu:,} 件')
    return pd.DataFrame(records, columns=['SC', 'date', 'type'])


# ---- メイン -----------------------------------------------------------------

def main():
    # 1. JPX xlsx を読み込む（kessan[0-9]*.xlsx）
    jpx_files = sorted(SETTLEMENT_DIR.glob('kessan[0-9]*.xlsx'))
    if not jpx_files:
        print('[警告] JPX xlsx が見つかりません（kessan[0-9]*.xlsx）')
    df_jpx = load_jpx_xlsx(jpx_files)

    # 2. kessan_pro xlsx を読み込む（kessan_pro_*.xlsx）
    kp_files = sorted(SETTLEMENT_DIR.glob('kessan_pro_*.xlsx'))
    if kp_files:
        df_kp = pd.concat(
            [load_kessan_pro(f) for f in kp_files],
            ignore_index=True,
        )
    else:
        print('[情報] kessan_pro_*.xlsx が見つかりません（スキップ）')
        df_kp = pd.DataFrame(columns=['SC', 'date', 'type'])

    # 3. マージ・重複除去（SC + date が同一の場合は JPX データを優先）
    df_all = pd.concat([df_jpx, df_kp], ignore_index=True)
    df_all['SC'] = df_all['SC'].astype(str).str.zfill(4)
    df_all = df_all.drop_duplicates(subset=['SC', 'date'], keep='first')
    df_all = df_all.sort_values(['date', 'SC']).reset_index(drop=True)

    # 4. UTF-8 で CSV 出力（ファイルがロック中の場合は _new サフィックスで保存）
    try:
        df_all.to_csv(OUTPUT_FILE, index=False, encoding='utf-8')
        out_path = OUTPUT_FILE
    except PermissionError:
        alt = OUTPUT_FILE.with_stem(OUTPUT_FILE.stem + '_new')
        df_all.to_csv(alt, index=False, encoding='utf-8')
        out_path = alt
        print(f'[警告] {OUTPUT_FILE.name} が別プロセスで開かれているため {alt.name} に保存しました')
        print(f'       Excel/Notepad を閉じてから {alt.name} を {OUTPUT_FILE.name} にリネームしてください')

    # 5. 確認: SC4885（室町ケミカル）が含まれているか
    sc4885 = df_all[df_all['SC'].astype(str) == '4885']
    if not sc4885.empty:
        print(f'\n[確認] SC4885 ({len(sc4885)} 件):')
        print(sc4885.to_string(index=False))
    else:
        print('\n[確認] SC4885: データなし')

    print(f'\n{out_path.name} に {len(df_all)} 件出力しました')
    print(f'  JPX 件数       : {len(df_jpx)}')
    print(f'  kessan_pro 件数: {len(df_kp)}')
    print(f'  重複除去後合計 : {len(df_all)}')
    print(f'出力先: {out_path}')


if __name__ == '__main__':
    main()
