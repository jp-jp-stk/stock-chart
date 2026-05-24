# build_master.py
# kessan_pro_*.xlsx から企業名マスタ（company_master.csv）を生成するスクリプト。
#
# 実行タイミング：
#   - kessan_pro_*.xlsx を新しいファイルに差し替えたとき
#   - 初回セットアップ時
#
# 使い方：
#   python build_master.py

from openpyxl import load_workbook
from pathlib import Path
import csv
import re

# ---- 設定 -------------------------------------------------------------------
SETTLEMENT_DIR = Path(r'C:\Users\jnpei\Documents\stock-chart\data\settlement')
OUTPUT_FILE    = SETTLEMENT_DIR / 'company_master.csv'
# -----------------------------------------------------------------------------


def normalize_sc(sc) -> str | None:
    """SC を 4 桁ゼロ埋め文字列に正規化する。"""
    if sc is None:
        return None
    if isinstance(sc, (int, float)):
        return str(int(sc)).zfill(4)
    s = str(sc).strip()
    return s.zfill(4) if s else None


def extract_company_name(raw) -> str | None:
    """=HYPERLINK("url","企業名") 形式のセルから企業名を抽出する。
    式がなければ str(raw) をそのまま返す。"""
    if raw is None:
        return None
    m = re.search(r',\"(.+?)\"', str(raw))
    return m.group(1) if m else str(raw)


def normalize_name(name: str) -> str:
    """全角英数字・全角スペースを半角に正規化する（検索用）。"""
    return name.translate(str.maketrans(
        '　ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ'
        'ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ'
        '０１２３４５６７８９',
        ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        'abcdefghijklmnopqrstuvwxyz'
        '0123456789',
    ))


def build_company_master(kp_files: list[Path]) -> int:
    """kessan_pro xlsx から SC→企業名マスタを生成し company_master.csv に出力する。

    フォーマット（0-indexed 列番号）:
      row[0] : SC コード
      row[1] : 企業名（=HYPERLINK 形式 or プレーンテキスト）
    """
    master: dict[str, str] = {}   # SC → 企業名

    for filepath in sorted(kp_files):
        print(f'読み込み中: {filepath.name}')
        try:
            # data_only=False で HYPERLINK 式を文字列として取得
            wb = load_workbook(filepath, read_only=True)
        except Exception as e:
            print(f'  スキップ（読み込みエラー）: {e}')
            continue

        ws        = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)   # ヘッダー行スキップ

        for row in rows_iter:
            if row is None or len(row) < 2:
                continue
            sc   = normalize_sc(row[0])
            name = extract_company_name(row[1])
            if sc and name and name not in ('None', '') and sc not in master:
                master[sc] = name

        wb.close()

    if not master:
        print('有効なデータが見つかりませんでした。')
        return 0

    SETTLEMENT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['SC', 'name', 'name_kana'])
        for sc, name in sorted(master.items()):
            writer.writerow([sc, name, normalize_name(name)])

    print(f'\ncompany_master.csv に {len(master)} 件出力しました')
    print(f'出力先: {OUTPUT_FILE}')
    return len(master)


def main() -> None:
    kp_files = sorted(SETTLEMENT_DIR.glob('kessan_pro_*.xlsx'))
    if not kp_files:
        print('kessan_pro_*.xlsx が見つかりません。')
        print(f'対象フォルダ: {SETTLEMENT_DIR}')
        return
    print(f'kessan_pro ファイル: {len(kp_files)} 件')
    build_company_master(kp_files)


if __name__ == '__main__':
    main()
